import csv
import os
from openpyxl import Workbook

def clean_path(p):
    return p.strip().strip("&").strip().strip("'").strip('"').replace("\\ ", " ").strip()

def read_csv(path):
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        return [row for row in reader], reader.fieldnames or []

def to_float(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return None

while True:
    path_a = clean_path(input("drop main file here: "))
    path_b = clean_path(input("drop second file here: "))

    try:
        rows_a, cols_a = read_csv(path_a)
        rows_b, cols_b = read_csv(path_b)
    except Exception as e:
        print(f"error reading files: {e}")
        continue

    has_src_a = 'SrcCr' in cols_a
    has_src_b = 'SrcCr' in cols_b
    has_cr_a  = 'Cr' in cols_a
    has_cr_b  = 'Cr' in cols_b

    if not has_src_a and not has_src_b:
        print("error: neither file looks like the main transfer file")
    elif not has_cr_a and not has_cr_b:
        print("error: neither file looks like the second transfer file")
    elif has_src_a and has_src_b:
        print("error: both files have SrcCr — drop different files")
    elif has_cr_a and has_cr_b:
        print("error: both files have Cr — drop different files")
    else:
        break

if has_src_a:
    rows1, cols1 = rows_a, cols_a
    rows2, cols2 = rows_b, cols_b
    main_path = path_a
else:
    rows1, cols1 = rows_b, cols_b
    rows2, cols2 = rows_a, cols_a
    main_path = path_b

# calc from main (transferred) file — skip rows where SrcCr is not numeric
valid = [r for r in rows1 if to_float(r.get('SrcCr')) is not None]

sub_total_initial_credit_1 = sum(to_float(r.get('SrcCr')) or 0 for r in rows1)
total_final_credit          = sum(to_float(r.get('TgtCr')) or 0 for r in valid)
initial_pre_GPA_1           = sum((to_float(r.get('Gr')) or 0) * (to_float(r.get('SrcCr')) or 0) for r in valid)
final_GPA                   = sum((to_float(r.get('Gr')) or 0) * (to_float(r.get('TgtCr')) or 0) for r in valid) / total_final_credit if total_final_credit else 0

# calc from second (non-transferred) file
sub_total_initial_credit_2 = sum(to_float(r.get('Cr')) or 0 for r in rows2)
initial_pre_GPA_2           = sum((to_float(r.get('Gr')) or 0) * (to_float(r.get('Cr')) or 0) for r in rows2)

# totals
total_initial_credit = sub_total_initial_credit_1 + sub_total_initial_credit_2
initial_GPA          = (initial_pre_GPA_1 + initial_pre_GPA_2) / total_initial_credit if total_initial_credit else 0

# write xlsx
wb = Workbook()

def write_sheet(wb, title, rows, cols, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.append(list(cols))
    for row in rows:
        ws.append([row.get(c, '') for c in cols])

write_sheet(wb, 'Transferred',     rows1, cols1, first=True)
write_sheet(wb, 'Non-Transferred', rows2, cols2)

ws3 = wb.create_sheet('Summary')
ws3.append(['Label', 'Value'])
ws3.append(['Initial GPA',     initial_GPA])
ws3.append(['Final GPA',       final_GPA])
ws3.append(['Initial Credits', total_initial_credit])
ws3.append(['Final Credits',   total_final_credit])

base_name   = os.path.splitext(os.path.basename(main_path))[0]
output_path = os.path.join(os.path.dirname(main_path), f'{base_name} - output.xlsx')
wb.save(output_path)

print(f"saved to: {output_path}")
input("press enter to close...")
